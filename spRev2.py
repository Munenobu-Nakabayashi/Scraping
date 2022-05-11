from json.tool import main
from bs4 import BeautifulSoup
import requests
import openpyxl     # 2022.05.10 ADD
import datetime

def main():

    today = datetime.date.today()
    yyyymmdd = today.strftime('%Y%m%d')

    wb = openpyxl.load_workbook('c:/Python/test/sp.xlsx')   # 2022.05.10 UPDATE

    try:
        wb.remove_sheet(wb.get_sheet_by_name(yyyymmdd))     # 本日日付シートが存在する場合は削除
    except:
        pass    # 無条件続行
            
    wb.create_sheet(index=0, title=yyyymmdd)
    wb.save('c:/Python/test/sp.xlsx')

    sheet = wb[yyyymmdd]

    """URLを取得"""
    gets = requests.get('http://jvn.jp/report')

    """bs4にてパーサーする（HTMLを公文表示させる）"""
    soup = BeautifulSoup(gets.text, 'html.parser')

    """要素の格納用に空のリストを準備"""
    airs = []

    """find_allでdtタグの要素を順番に全て取得し空リストに格納"""

    for tag in soup.find_all('dt'):
        airs.append(tag.text)
    
    i = 1
    for air in airs:

        work = air.strip()
        wk = work.replace('ã', '')    # 全角スペース文字化け部分を除去
        sheet.cell(row=i, column=1, value=wk[0:10])
        sheet.cell(row=i, column=2, value=wk[10:].replace(':', ''))
        # print(wk[10:15])
        if str(wk[10:15]) == "JVNVU":
            urls = "http://jvn.jp/vu/" + wk[10:].replace(':', '').replace('#', '') + "/index.html"
        else:
            urls = "http://jvn.jp/jp/" + wk[10:].replace(':', '').replace('#', '') + "/index.html"
        sheet.cell(row=i, column=3, value=urls)

        i = i + 1

    wb.save('c:/Python/test/sp.xlsx')

main()  # 2022.05.10 UPDATE