#!/usr/bin/python
# -*- coding: utf-8 -*-
from json.tool import main
from bs4 import BeautifulSoup
import requests
import openpyxl     # 2022.05.10 ADD
import datetime

import sqlite3      # 2022.05.16 UPDATE

def main():

    today = datetime.date.today()
    yyyymmdd = today.strftime('%Y%m%d')

    wb = openpyxl.load_workbook('c:/Python/test/sp.xlsx')   # 2022.05.10 UPDATE

    try:
        wb.remove_sheet(wb.get_sheet_by_name(yyyymmdd)) 
    except:
        pass
            
    wb.create_sheet(index=0, title=yyyymmdd)
    wb.save('c:/Python/test/sp.xlsx')

    sheet = wb[yyyymmdd]

    db_path = "c:\Python\SQLite\jvn.db"     # 2022.05.16 UPDATE
    conn = sqlite3.connect(db_path)         # 2022.05.16 UPDATE
    cur = conn.cursor()

    gets = requests.get('http://jvn.jp/report')


    soup = BeautifulSoup(gets.text, 'html.parser')


    airs = []

    for tag in soup.find_all('dt'):
        airs.append(tag.text)
    
    i = 1
    for air in airs:

        work = air.strip()
        wk = work.replace('ã', '')    
        sheet.cell(row=i, column=1, value=wk[0:10])
        sheet.cell(row=i, column=2, value=wk[10:].replace(':', ''))
        # print(wk[10:15])
        if str(wk[10:15]) == "JVNVU":
            urls = "http://jvn.jp/vu/" + wk[10:].replace(':', '').replace('#', '') + "/index.html"
        else:
            urls = "http://jvn.jp/jp/" + wk[10:].replace(':', '').replace('#', '') + "/index.html"
        sheet.cell(row=i, column=3, value=urls)

        # ADD 2022.05.16 --- START
        jvndate = wk[0:10]
        jvncode = wk[10:].replace(':', '')
        jvnurl = urls

        ### sql = 'INSERT INTO JVN (JVNCODE, JVNDATE, JVNURL) values (?,?,?)'
        sql = 'INSERT INTO JVN (JVNCODE, JVNDATE, JVNURL) values (?,?,?) ON CONFLICT DO NOTHING'
        # print(jvncode)
        # print(jvndate)
        # print(jvnurl)
        data = [jvncode, jvndate, jvnurl]
        cur.execute(sql, data)
        # ADD 2022.05.16 --- END

        i = i + 1

    wb.save('c:/Python/test/sp.xlsx')

    conn.commit()       # 2022.05.16 UPDATE
    conn.close()        # 2022.05.16 UPDATE

main()  # 2022.05.10 UPDATE